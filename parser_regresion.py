"""Parser del Excel de regresión (26_08_21_-_Regresion.xlsx).

Espera una hoja con columnas:
    Año | Mes | Yt | X1 | X2 | X3 | X4

Y en columnas laterales los coeficientes:
    X1 → b1
    X2 → b2
    X3 → b3
    X4 → b4
    MAPE
    R2
"""

import openpyxl


def parse_regresion_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # Leer coeficientes (buscamos en las primeras filas, columnas I/J)
    coefs = {'b1': None, 'b2': None, 'b3': None, 'b4': None, 'mape': None, 'r2': None}
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=12, values_only=True):
        label = str(row[8]).strip().upper() if row[8] is not None else ''
        val = row[9]
        if label == 'X1' and val is not None:
            coefs['b1'] = float(val)
        elif label == 'X2' and val is not None:
            coefs['b2'] = float(val)
        elif label == 'X3' and val is not None:
            coefs['b3'] = float(val)
        elif label == 'X4' and val is not None:
            coefs['b4'] = float(val)
        elif label == 'MAPE' and val is not None:
            coefs['mape'] = float(val)
        elif label == 'R2' and val is not None:
            coefs['r2'] = float(val)

    if any(coefs[k] is None for k in ('b1', 'b2', 'b3', 'b4')):
        raise ValueError(
            f"No se pudieron leer todos los coeficientes. Encontrados: {coefs}"
        )

    # Leer puntos de datos
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else '' for c in row[:7]]
        if 'año' in cells or 'anio' in cells or 'year' in cells:
            header_row = i
            break
    if header_row is None:
        header_row = 1

    points = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=7, values_only=True):
        if row[0] is None:
            continue
        try:
            anio = int(row[0])
            mes = int(row[1])
            yt = float(row[2]) if row[2] is not None else None
            x1 = float(row[3])
            x2 = float(row[4])
            x3 = float(row[5])
            x4 = float(row[6])
            points.append({
                'anio': anio,
                'mes': mes,
                'yt': yt if yt and yt > 0 else None,
                'x1': x1,
                'x2': x2,
                'x3': x3,
                'x4': x4,
            })
        except (TypeError, ValueError):
            continue

    return coefs, points
