"""Parser de la planilla de VOLUMEN, en Excel o CSV.

Las columnas se leen SIEMPRE por posición, sin mirar los encabezados:

    A=Año  B=Mes  C=Petrolera  D=provincia  E=Sector  F=producto  G=Volumen

Los encabezados suelen venir con la codificación rota (ej. "AÃ±o" en vez de
"Año") cuando la planilla pasó por Excel con la configuración regional
equivocada, así que matchear por nombre no es confiable. La primera fila se
saltea salvo que su columna A ya contenga un año válido, para no perder una
fila de datos si la planilla viene sin encabezado.

Volumen admite coma o punto decimal y separadores de miles.
"""

import csv
import io
import re
from collections import defaultdict

from openpyxl import load_workbook

# Posición de cada campo: la planilla siempre trae estas columnas en este orden.
COL_ANIO, COL_MES, COL_PETROLERA, COL_PROVINCIA, COL_SECTOR, COL_PRODUCTO, COL_VOLUMEN = range(7)
COLUMNAS_MINIMAS = 7

PRODUCTOS_VALIDOS = ('GO2', 'GO3', 'N2', 'N3')
ANIO_MIN, ANIO_MAX = 1990, 2100


# Normalización de nombres de provincia
PROVINCE_ALIASES = {
    'capital federal': 'CABA',
    'caba': 'CABA',
    'ciudad autonoma de buenos aires': 'CABA',
    'entre rios': 'Entre Ríos',
    'entre ríos': 'Entre Ríos',
    'rio negro': 'Río Negro',
    'río negro': 'Río Negro',
    'tucuman': 'Tucumán',
    'tucumán': 'Tucumán',
    'neuquen': 'Neuquén',
    'córdoba': 'Córdoba',
    'cordoba': 'Córdoba',
    'santa fe': 'Santa Fe',
    'tierra del fuego': 'Tierra del Fuego',
    'santiago del estero': 'Santiago del Estero',
    'la pampa': 'La Pampa',
    'la rioja': 'La Rioja',
    'san juan': 'San Juan',
    'san luis': 'San Luis',
    'santa cruz': 'Santa Cruz',
    'buenos aires': 'Buenos Aires',
    'estado nacional': 'Estado Nacional',
    'no aplica': 'No aplica',
    'provincia': 'No aplica',
}

CANONICAL_PROVINCES = [
    'Buenos Aires', 'CABA', 'Catamarca', 'Chaco', 'Chubut', 'Córdoba',
    'Corrientes', 'Entre Ríos', 'Formosa', 'Jujuy', 'La Pampa', 'La Rioja',
    'Mendoza', 'Misiones', 'Neuquén', 'Río Negro', 'Salta', 'San Juan',
    'San Luis', 'Santa Cruz', 'Santa Fe', 'Santiago del Estero',
    'Tierra del Fuego', 'Tucumán',
]


def normalize_province(name: str) -> str:
    if not name:
        return 'No aplica'
    key = name.strip().lower()
    key_ascii = (key
                 .replace('á', 'a').replace('é', 'e').replace('í', 'i')
                 .replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n'))
    for alias, canon in PROVINCE_ALIASES.items():
        alias_ascii = (alias
                       .replace('á', 'a').replace('é', 'e').replace('í', 'i')
                       .replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n'))
        if key == alias or key_ascii == alias_ascii:
            return canon
    return name.strip().title()


def parse_volumen(raw: str) -> float:
    if raw is None:
        return 0.0
    s = str(raw).strip()
    if not s or s in ('-', 'S/N', 'null', 'None'):
        return 0.0
    s = s.replace(' ', '')
    if s.count('.') > 1:
        parts = s.split('.')
        if len(parts[-1]) <= 2:
            s = ''.join(parts[:-1]) + '.' + parts[-1]
        else:
            s = ''.join(parts)
    s = s.replace(',', '.')
    try:
        val = float(s)
    except ValueError:
        return 0.0
    if abs(val) > 50_000_000:
        return 0.0
    return val


def reparar_texto(valor):
    """Repara texto UTF-8 leído como latin-1 ("Al PÃºblico" -> "Al Público").

    Pasa cada vez que la planilla se abre en Excel con la configuración
    regional equivocada. Sin esto quedarían dos sectores distintos para el
    mismo nombre. Solo se aplica si el texto tiene la firma del problema y la
    reconversión no falla.
    """
    if not isinstance(valor, str) or ('Ã' not in valor and 'Â' not in valor):
        return valor
    try:
        return valor.encode('latin-1').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        return valor


def _fila(valores):
    """Convierte una fila posicional en el dict que espera la base.

    Devuelve None si la fila no sirve: encabezado, mes fuera de rango,
    producto desconocido o año ilegible.
    """
    if len(valores) < COLUMNAS_MINIMAS:
        return None
    try:
        anio = int(float(str(valores[COL_ANIO]).strip()))
        mes = int(float(str(valores[COL_MES]).strip()))
    except (TypeError, ValueError, AttributeError):
        return None
    if not (ANIO_MIN <= anio <= ANIO_MAX) or not (1 <= mes <= 12):
        return None
    producto = str(valores[COL_PRODUCTO] or '').strip().upper()
    if producto not in PRODUCTOS_VALIDOS:
        return None
    return {
        'anio': anio,
        'mes': mes,
        'petrolera': reparar_texto(str(valores[COL_PETROLERA] or '').strip()),
        'provincia': normalize_province(reparar_texto(str(valores[COL_PROVINCIA] or ''))),
        'sector': reparar_texto(str(valores[COL_SECTOR] or '').strip()) or 'S/N',
        'producto': producto,
        'volumen': parse_volumen(valores[COL_VOLUMEN]),
    }


def _procesar(filas):
    """Recorre filas posicionales, salteando el encabezado si lo hay."""
    rows, skipped = [], 0
    for i, valores in enumerate(filas):
        fila = _fila(list(valores))
        if fila is None:
            if i > 0:          # la fila 1 suele ser el encabezado
                skipped += 1
            continue
        rows.append(fila)
    if rows:
        anios = [r['anio'] for r in rows]
        a0, a1 = min(anios), max(anios)
        periodo = '%d-%02d a %d-%02d' % (
            a0, min(r['mes'] for r in rows if r['anio'] == a0),
            a1, max(r['mes'] for r in rows if r['anio'] == a1))
    else:
        periodo = '—'
    info = {
        'volumen_max': max((r['volumen'] for r in rows), default=0.0),
        'volumen_total': sum(r['volumen'] for r in rows),
        'sectores': sorted({r['sector'] for r in rows}),
        'petroleras': sorted({r['petrolera'] for r in rows}),
        'periodo': periodo,
    }
    return rows, skipped, info


def parse_excel(file_obj):
    """Lee la primera hoja de un .xlsx/.xlsm por posición de columna."""
    datos = file_obj.read()
    if not isinstance(datos, bytes):
        datos = datos.encode('utf-8')
    wb = load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        return _procesar(ws.iter_rows(min_col=1, max_col=COLUMNAS_MINIMAS,
                                      values_only=True))
    finally:
        wb.close()


def parse_csv(file_obj, encoding='utf-8'):
    """Lee un CSV separado por ; por posición de columna."""
    content = file_obj.read()
    if isinstance(content, bytes):
        for enc in (encoding, 'utf-8-sig', 'latin-1', 'cp1252'):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode('utf-8', errors='replace')
    else:
        text = content
    if text.startswith('\ufeff'):
        text = text[1:]
    return _procesar(csv.reader(text.splitlines(), delimiter=';'))


def parse_archivo(file_obj, filename=''):
    """Elige el lector segun la extension del archivo subido."""
    nombre = (filename or '').lower()
    if nombre.endswith(('.xlsx', '.xlsm', '.xltx')):
        return parse_excel(file_obj)
    if nombre.endswith('.xls'):
        raise ValueError(
            'El formato .xls antiguo no se puede leer. Guardá la planilla '
            'como .xlsx desde Excel y volvé a subirla.')
    return parse_csv(file_obj)
